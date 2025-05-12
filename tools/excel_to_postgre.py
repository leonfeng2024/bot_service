import sys
import openpyxl
import re
import os
from sqlalchemy import create_engine, inspect, text


def extract_dataset_name(file_path):
    """
    Extract dataset name from file name
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        str: Dataset name
    """
    _file_name_list = os.path.basename(file_path).split(".")[0].split("_")
    if len(_file_name_list) >= 3:
        dataset_name = _file_name_list[0] + "_" + _file_name_list[2]
    else:
        # Fallback if file name doesn't match expected format
        dataset_name = os.path.basename(file_path).split(".")[0]
    return dataset_name


def extract_tables_from_excel(file_path, sheet_name="資材一覧", header_row=2, table_col_name="テーブル名"):
    """
    Extract table names from Excel file
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet containing table info
        header_row: Row number containing headers (1-based)
        table_col_name: Column name containing table names
        
    Returns:
        set: Set of table names
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        # Try to get the specified sheet, or use the first sheet if not found
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            
        header = [cell.value for cell in ws[header_row]]

        if table_col_name in header:
            table_col_idx = header.index(table_col_name) + 1
        else:
            print(f"Column '{table_col_name}' not found in sheet")
            return set()

        filtered = set()
        for row in range(header_row + 1, ws.max_row + 1):
            table_cell = ws.cell(row=row, column=table_col_idx)

            if not (table_cell.font and table_cell.font.strike):
                if table_cell.value is not None:
                    filtered.add(table_cell.value)

        # Filter only values containing at least one letter
        filtered_set = {s for s in filtered if re.search(r"[a-zA-Z]", s)}
        return filtered_set
        
    except Exception as e:
        print(f"Error extracting tables from Excel: {str(e)}")
        return set()


def extract_table_mappings(file_path, sheet_name="テーブル一覧"):
    """
    Extract table mappings from the テーブル一覧 sheet
    
    Args:
        file_path: Path to the Excel file
        sheet_name: Name of the sheet containing the table mappings
        
    Returns:
        dict: Mapping of table physical names to logical names
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found in Excel file")
            return {}
            
        ws = wb[sheet_name]
        
        # Table structure:
        # We need to extract the columns for physical table names and logical table names
        
        # First, locate the header row
        header_row = None
        physical_name_col = None
        logical_name_col = None
        
        # Look for headers within the first 10 rows
        for row in range(1, min(10, ws.max_row + 1)):
            for col in range(1, min(10, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value == "テーブル物理名":  # Physical table name
                    physical_name_col = col
                    header_row = row
                elif cell_value == "テーブル論理名":  # Logical table name
                    logical_name_col = col
        
        if not header_row or not physical_name_col or not logical_name_col:
            print(f"Could not find required headers in sheet {sheet_name}")
            print(f"Headers found: header_row={header_row}, physical_name_col={physical_name_col}, logical_name_col={logical_name_col}")
            return {}
        
        print(f"Found headers at row {header_row}: Physical table name at column {physical_name_col}, Logical table name at column {logical_name_col}")
        
        mappings = {}
        # Start reading from the row after the header
        for row in range(header_row + 1, ws.max_row + 1):
            physical_name = ws.cell(row=row, column=physical_name_col).value
            logical_name = ws.cell(row=row, column=logical_name_col).value
            
            if physical_name and logical_name:
                # Process physical and logical names as strings, ensuring whitespace is removed
                physical_name = str(physical_name).strip()
                logical_name = str(logical_name).strip()
                
                # Exclude formulas or special values
                if not physical_name.startswith("=") and not logical_name.startswith("="):
                    mappings[physical_name] = logical_name
                    print(f"Mapped: {physical_name} -> {logical_name}")
        
        print(f"Table mappings found: {len(mappings)} entries")
        return mappings
        
    except Exception as e:
        print(f"Error extracting table mappings from Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        return {}


def create_required_tables(db_uri):
    """
    Create necessary tables and views if they don't exist
    
    Args:
        db_uri: Database URI
    """
    engine = create_engine(db_uri)
    inspector = inspect(engine)
    all_tables = inspector.get_table_names()
    all_views = inspector.get_view_names()

    not_in_tables = [item for item in ["dataset_view_tables"] if item not in all_tables]
    not_in_views = [item for item in ["v_dataset_view_table_field"] if item not in all_views]

    create_sql = []
    if len(not_in_tables) > 0:
        if "dataset_view_tables" in not_in_tables:
            create_sql.append("""
                CREATE TABLE dataset_view_tables (
                    physical_name TEXT NOT NULL,
                    logical_name TEXT,
                    nm TEXT NOT NULL,
                    nm_jpn TEXT
                );""")

    if len(not_in_views) > 0:
        if "v_dataset_view_table_field" in not_in_views:
            create_sql.append("""
                create view v_dataset_view_table_field as
                select dvf.physical_name as ds_physical_name, dvf.logical_name as ds_logical_name, dvf.nm as view_name, alv.table_physical_name as table_name, alv.field as field, alv.field_jpn as field_jpn 
                from dataset_view_tables as dvf
                left join 
                (select * from v_view_table_field
                union all 
                select tf.physical_name as view_physical_name, tf.logical_name as view_logical_name, tf.physical_name as table_physical_name, tf.logical_name as table_logical_name, tf.field as field, tf.field_jpn as field_jpn 
                from table_fields as tf) as alv on dvf.nm=alv.view_physical_name
                order by dvf.physical_name, dvf.nm, alv.field;
            """)

    if len(create_sql) > 0:
        with engine.connect() as conn:
            for s in create_sql:
                conn.execute(text(s))
            conn.commit()


def insert_dataset_tables(db_uri, dataset_name, table_set, table_mappings=None):
    """
    Insert dataset-table relationships into PostgreSQL
    
    Args:
        db_uri: Database URI
        dataset_name: Dataset name
        table_set: Set of table names
        table_mappings: Dictionary mapping table physical names to logical names
    """
    if not table_set:
        print("No tables to insert")
        return
        
    if table_mappings is None:
        table_mappings = {}
        
    engine = create_engine(db_uri)
    with engine.connect() as conn:
        # delete the old data
        delete_sql = "DELETE FROM dataset_view_tables WHERE physical_name = :physical_name"
        conn.execute(statement=text(delete_sql), parameters={"physical_name": dataset_name})
        
        # insert new data with logical_name and nm_jpn
        insert_sql = """
            INSERT INTO dataset_view_tables 
            (physical_name, logical_name, nm, nm_jpn) 
            VALUES (:physical_name, :logical_name, :nm, :nm_jpn)
        """
        for table_name in table_set:
            # Set logical_name to same as physical_name as required
            logical_name = dataset_name
            # Get nm_jpn from table_mappings if available
            nm_jpn = table_mappings.get(table_name)
            
            conn.execute(
                statement=text(insert_sql), 
                parameters={
                    "physical_name": dataset_name, 
                    "logical_name": logical_name,
                    "nm": table_name,
                    "nm_jpn": nm_jpn
                }
            )

        conn.commit()


def process_excel_file(file_path, db_uri, sheet_name="資材一覧", header_row=2, table_col_name="テーブル名"):
    """
    Process Excel file and import table data to PostgreSQL
    
    Args:
        file_path: Path to the Excel file
        db_uri: Database URI
        sheet_name: Name of the sheet containing table info
        header_row: Row number containing headers (1-based)
        table_col_name: Column name containing table names
        
    Returns:
        tuple: (dataset_name, table_set) - The extracted dataset name and table set
    """
    try:
        # Extract dataset name from file name
        dataset_name = extract_dataset_name(file_path)
        
        # Extract table names from Excel
        table_set = extract_tables_from_excel(file_path, sheet_name, header_row, table_col_name)
        
        # Extract table mappings from the table list sheet
        table_mappings = extract_table_mappings(file_path)
        
        # Create required tables and views
        create_required_tables(db_uri)
        
        # Insert dataset-table relationships with mappings
        insert_dataset_tables(db_uri, dataset_name, table_set, table_mappings)
        
        return dataset_name, table_set
        
    except Exception as e:
        print(f"Error processing Excel file: {str(e)}")
        raise


# Example usage (will only run if script is executed directly)
if __name__ == "__main__":
    file_path = r"/temp/PD011_02納入_納入計画実績(MR向け).xlsx"
    db_uri = "postgresql+psycopg2://postgres:123456@localhost:5432/verify5"
    
    dataset_name, table_set = process_excel_file(file_path, db_uri)
    
    print(f"Dataset name: {dataset_name}")
    print(f"Tables found: {table_set}")

